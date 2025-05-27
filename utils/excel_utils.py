import openpyxl
from config.config import *

def read_excel(file_path=EXCEL_FILE, sheet_name=SHEET_NAME):
    """
    读取excel：
    使用openpyxl库里的方法加载excel 并且选择excel里的表
	然后读取excel第二行的关键字数据存到list
	然后遍历excel第三行开始的数据读取到元组tuple
	用zip函数把他们两个转化成键值对
	用dict把它变成字典
	最后判断其中的istrue关键字是true的话就添加到字典列表data里面
	关闭文件 返回字典列表data)"""

    """
    
    打开 excel 文件
    使用 openpyxl.load_workbook(file_path) 加载 Excel 文件，创建 workbook 对象
	openpyxl 是 Python 处理 Excel .xlsx 文件的库
    """
    workbook = openpyxl.load_workbook(file_path)   # 参数传文件路径
    # workbook = openpyxl.load_workbook("../data/测试用例.xlsx")   # 参数传文件路径
    """选择excel里指定的表sheet"""
    worksheet = workbook[sheet_name]
    """data字典列表：最终存储 Excel 数据的列表，每一行 Excel 数据都会转换成一个dict存入data"""
    data = []
    """
    拿key行, 也就是表的第二行, 生成一个 key 的列表，这行数据用于作为字典的 key
	    [cell.value for cell in worksheet[2]] 遍历第二行的所有单元格 提取每个单元格的值
	"""
    keys = [cell.value for cell in worksheet[2]]
    # print("keys=", keys, "\ntype-keys=", type(keys))
    """
    从第 3 行开始，逐行读取 Excel 数据，每一行返回一个元组
	    values_only=True 让 row 只包含值，不包含 Cell 对象"""
    for row in worksheet.iter_rows(min_row=3, values_only=True):
        """
        zip(keys, row) 把 keys 和 row 组合成键值对
	•	dict(zip(keys, row)) 转换为字典
	    """
        # print("row=", row, "\ntype-row=", type(row))
        dict_data = dict(zip(keys, row))
        # print("dict_data=", dict_data, "\ntype-dict_data=", type(dict_data))
        """如果 dict_data["is_true"] 为 True，则把数据添加到 data 列表中；为 False 或 None，则跳过该行数据"""
        if dict_data["is_true"]:
            data.append(dict_data)
        # data.append(dict_data)
    # print(data) # 打印拿到的所有数据
    # 关闭 excel 文件
    workbook.close()
    # print("data=",data,"\ntype-date=",type(data))
    return data
"""
    keys = ["id", "name", "is_true"]
	row = (1, "张三", True)
	dict_data = {"id": 1, "name": "张三", "is_true": True}
    data=
    [
        {"id": 1, "name": "张三", "is_true": True},
        {"id": 3, "name": "王五", "is_true": True}
    ]
"""
# read_excel()
# if __name__ == "__main__":
#     read_excel()